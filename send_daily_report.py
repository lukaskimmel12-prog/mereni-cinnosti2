from __future__ import annotations

import os
import smtplib
from datetime import datetime, timedelta, timezone
from email.message import EmailMessage
from io import BytesIO
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from supabase import create_client


# ============================================================
# NASTAVENÍ
# ============================================================

APP_TIMEZONE = ZoneInfo("Europe/Prague")

ACTIVITIES = [
    "Aperam",
    "Personna",
    "SSI",
    "Zanini",
    "Rebound",
]

YUSEN_BLUE = "00529B"
YUSEN_ORANGE = "F58220"
WHITE = "FFFFFF"
LIGHT_BLUE = "DDEBF7"


# ============================================================
# GITHUB SECRETS
# ============================================================

def get_required_secret(name: str) -> str:
    value = os.getenv(name)

    if not value:
        raise RuntimeError(
            f"V GitHub Secrets chybí hodnota: {name}"
        )

    return value


SUPABASE_URL = get_required_secret("SUPABASE_URL")
SUPABASE_KEY = get_required_secret("SUPABASE_KEY")
GMAIL_ADDRESS = get_required_secret("GMAIL_ADDRESS")
GMAIL_APP_PASSWORD = get_required_secret(
    "GMAIL_APP_PASSWORD"
)
REPORT_RECIPIENT = get_required_secret(
    "REPORT_RECIPIENT"
)


# ============================================================
# DATUM A ČAS
# ============================================================

def parse_datetime(value: str) -> datetime:
    return datetime.fromisoformat(
        value.replace("Z", "+00:00")
    )


def to_local_datetime(value: str) -> datetime:
    return parse_datetime(value).astimezone(
        APP_TIMEZONE
    )


def format_duration(
    seconds: int | float | None,
) -> str:
    total_seconds = max(0, int(seconds or 0))

    hours, remainder = divmod(
        total_seconds,
        3600,
    )

    minutes, seconds = divmod(
        remainder,
        60,
    )

    return (
        f"{hours:02d}:"
        f"{minutes:02d}:"
        f"{seconds:02d}"
    )


def get_record_duration(
    record: dict,
    current_utc: datetime,
) -> int:
    saved_duration = record.get(
        "duration_seconds"
    )

    if saved_duration is not None:
        return max(0, int(saved_duration))

    start_time = parse_datetime(
        record["start_time"]
    )

    return max(
        0,
        int(
            (
                current_utc - start_time
            ).total_seconds()
        ),
    )


# ============================================================
# NAČTENÍ ZÁZNAMŮ ZE SUPABASE
# ============================================================

def load_records() -> tuple[
    list[dict],
    datetime,
    datetime,
]:
    database = create_client(
        SUPABASE_URL,
        SUPABASE_KEY,
    )

    period_end_utc = datetime.now(timezone.utc)
    period_start_utc = (
        period_end_utc
        - timedelta(hours=24)
    )

    response = (
        database.table("activity_log")
        .select("*")
        .gte(
            "start_time",
            period_start_utc.isoformat(),
        )
        .lt(
            "start_time",
            period_end_utc.isoformat(),
        )
        .order(
            "start_time",
            desc=False,
        )
        .execute()
    )

    return (
        response.data or [],
        period_start_utc,
        period_end_utc,
    )


# ============================================================
# SOUHRN
# ============================================================

def calculate_activity_totals(
    records: list[dict],
    current_utc: datetime,
) -> dict[str, int]:
    totals = {
        activity: 0
        for activity in ACTIVITIES
    }

    for record in records:
        activity = record.get(
            "activity",
            "Neznámá",
        )

        duration = get_record_duration(
            record,
            current_utc,
        )

        if activity not in totals:
            totals[activity] = 0

        totals[activity] += duration

    return totals


# ============================================================
# FORMÁTOVÁNÍ EXCELU
# ============================================================

def style_header_row(
    worksheet,
    row_number: int = 1,
) -> None:
    fill = PatternFill(
        fill_type="solid",
        fgColor=YUSEN_BLUE,
    )

    font = Font(
        color=WHITE,
        bold=True,
    )

    for cell in worksheet[row_number]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )


def set_column_widths(
    worksheet,
    widths: dict[int, float],
) -> None:
    for column_number, width in widths.items():
        column_letter = get_column_letter(
            column_number
        )

        worksheet.column_dimensions[
            column_letter
        ].width = width


# ============================================================
# VYTVOŘENÍ EXCELU
# ============================================================

def create_excel(
    records: list[dict],
    period_end_utc: datetime,
) -> tuple[bytes, str]:
    workbook = Workbook()

    detail_sheet = workbook.active
    detail_sheet.title = "Detail"

    summary_sheet = workbook.create_sheet(
        "Souhrn"
    )

    # --------------------------------------------------------
    # LIST DETAIL
    # --------------------------------------------------------

    detail_headers = [
        "Datum",
        "ID",
        "Jméno",
        "Činnost",
        "Start",
        "Konec",
        "Trvání",
        "Trvání v minutách",
        "Stav",
    ]

    detail_sheet.append(detail_headers)

    for record in records:
        start_local = to_local_datetime(
            record["start_time"]
        )

        end_value = record.get("end_time")

        end_local = (
            to_local_datetime(end_value)
            if end_value
            else None
        )

        duration_seconds = get_record_duration(
            record,
            period_end_utc,
        )

        detail_sheet.append(
            [
                start_local.strftime(
                    "%d.%m.%Y"
                ),
                record.get(
                    "employee_id",
                    "",
                ),
                record.get(
                    "employee_name",
                    "",
                ),
                record.get(
                    "activity",
                    "",
                ),
                start_local.strftime(
                    "%H:%M:%S"
                ),
                (
                    end_local.strftime(
                        "%H:%M:%S"
                    )
                    if end_local
                    else ""
                ),
                format_duration(
                    duration_seconds
                ),
                round(
                    duration_seconds / 60,
                    2,
                ),
                (
                    "Dokončeno"
                    if end_value
                    else "Probíhá"
                ),
            ]
        )

    style_header_row(detail_sheet)

    detail_sheet.freeze_panes = "A2"
    detail_sheet.auto_filter.ref = (
        detail_sheet.dimensions
    )

    set_column_widths(
        detail_sheet,
        {
            1: 14,
            2: 12,
            3: 28,
            4: 18,
            5: 12,
            6: 12,
            7: 14,
            8: 21,
            9: 14,
        },
    )

    # --------------------------------------------------------
    # LIST SOUHRN
    # --------------------------------------------------------

    totals = calculate_activity_totals(
        records,
        period_end_utc,
    )

    total_seconds = sum(totals.values())

    summary_sheet.append(
        [
            "Činnost",
            "Trvání",
            "Minuty",
            "Podíl",
        ]
    )

    for activity in ACTIVITIES:
        seconds = totals.get(activity, 0)

        percentage = (
            seconds / total_seconds
            if total_seconds > 0
            else 0
        )

        summary_sheet.append(
            [
                activity,
                format_duration(seconds),
                round(seconds / 60, 2),
                percentage,
            ]
        )

    # Případné další neznámé činnosti
    for activity, seconds in totals.items():
        if activity in ACTIVITIES:
            continue

        percentage = (
            seconds / total_seconds
            if total_seconds > 0
            else 0
        )

        summary_sheet.append(
            [
                activity,
                format_duration(seconds),
                round(seconds / 60, 2),
                percentage,
            ]
        )

    style_header_row(summary_sheet)

    summary_sheet.freeze_panes = "A2"
    summary_sheet.auto_filter.ref = (
        summary_sheet.dimensions
    )

    set_column_widths(
        summary_sheet,
        {
            1: 20,
            2: 16,
            3: 16,
            4: 14,
        },
    )

    last_summary_row = (
        summary_sheet.max_row
    )

    for row_number in range(
        2,
        last_summary_row + 1,
    ):
        summary_sheet[
            f"D{row_number}"
        ].number_format = "0.0%"

    # Celkem
    total_row = last_summary_row + 2

    summary_sheet[
        f"A{total_row}"
    ] = "CELKEM"

    summary_sheet[
        f"B{total_row}"
    ] = format_duration(total_seconds)

    summary_sheet[
        f"C{total_row}"
    ] = round(total_seconds / 60, 2)

    for cell in summary_sheet[total_row]:
        cell.font = Font(
            bold=True,
            color=YUSEN_BLUE,
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=LIGHT_BLUE,
        )

    # --------------------------------------------------------
    # KOLÁČOVÝ GRAF
    # --------------------------------------------------------

    chart = PieChart()

    chart.title = (
        "Rozdělení času podle činností"
    )

    chart.height = 11
    chart.width = 16

    labels = Reference(
        summary_sheet,
        min_col=1,
        min_row=2,
        max_row=last_summary_row,
    )

    values = Reference(
        summary_sheet,
        min_col=3,
        min_row=1,
        max_row=last_summary_row,
    )

    chart.add_data(
        values,
        titles_from_data=True,
    )

    chart.set_categories(labels)

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showLeaderLines = True
    chart.dataLabels.showLegendKey = False

    summary_sheet.add_chart(
        chart,
        "F2",
    )

    # Zvýraznění názvu souhrnu
    summary_sheet["F15"] = (
        "Procenta jsou vypočítána "
        "z celkového času všech činností."
    )

    summary_sheet["F15"].font = Font(
        italic=True,
        color=YUSEN_BLUE,
    )

    # --------------------------------------------------------
    # ULOŽENÍ DO PAMĚTI
    # --------------------------------------------------------

    output = BytesIO()
    workbook.save(output)

    current_local = datetime.now(
        APP_TIMEZONE
    )

    filename = (
        "prehled_cinnosti_"
        + current_local.strftime(
            "%Y-%m-%d_%H-%M"
        )
        + ".xlsx"
    )

    return output.getvalue(), filename


# ============================================================
# TEXT SOUHRNU DO E-MAILU
# ============================================================

def create_email_summary(
    records: list[dict],
    period_end_utc: datetime,
) -> str:
    totals = calculate_activity_totals(
        records,
        period_end_utc,
    )

    total_seconds = sum(totals.values())

    lines = [
        "Souhrn podle činností:",
        "",
    ]

    for activity in ACTIVITIES:
        seconds = totals.get(activity, 0)

        percentage = (
            seconds / total_seconds * 100
            if total_seconds > 0
            else 0
        )

        lines.append(
            f"{activity}: "
            f"{percentage:.1f} % "
            f"({format_duration(seconds)})"
        )

    lines.extend(
        [
            "",
            (
                "Celkový zaznamenaný čas: "
                f"{format_duration(total_seconds)}"
            ),
        ]
    )

    return "\n".join(lines)


# ============================================================
# ODESLÁNÍ E-MAILU
# ============================================================

def send_email(
    excel_data: bytes,
    filename: str,
    records: list[dict],
    period_start_utc: datetime,
    period_end_utc: datetime,
) -> None:
    period_start_local = (
        period_start_utc.astimezone(
            APP_TIMEZONE
        )
    )

    period_end_local = (
        period_end_utc.astimezone(
            APP_TIMEZONE
        )
    )

    summary_text = create_email_summary(
        records,
        period_end_utc,
    )

    message = EmailMessage()

    message["From"] = GMAIL_ADDRESS
    message["To"] = REPORT_RECIPIENT
    message["Subject"] = (
        "Denní přehled činností UWH – "
        + period_end_local.strftime(
            "%d.%m.%Y"
        )
    )

    message.set_content(
        f"""Dobrý den,

v příloze zasílám automatický přehled činností za posledních 24 hodin.

Období:
{period_start_local.strftime("%d.%m.%Y %H:%M")}
až
{period_end_local.strftime("%d.%m.%Y %H:%M")}

Počet záznamů: {len(records)}

{summary_text}

Excel obsahuje dvě záložky:
- Detail – všechny jednotlivé záznamy,
- Souhrn – procentuální přehled a koláčový graf.

Tento e-mail byl vytvořen automaticky.
"""
    )

    message.add_attachment(
        excel_data,
        maintype="application",
        subtype=(
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        filename=filename,
    )

    with smtplib.SMTP_SSL(
        "smtp.gmail.com",
        465,
        timeout=30,
    ) as smtp:
        smtp.login(
            GMAIL_ADDRESS,
            GMAIL_APP_PASSWORD,
        )

        smtp.send_message(message)


# ============================================================
# SPUŠTĚNÍ
# ============================================================

def main() -> None:
    print(
        "Spouštím nový report s listy "
        "Detail a Souhrn."
    )

    (
        records,
        period_start_utc,
        period_end_utc,
    ) = load_records()

    excel_data, filename = create_excel(
        records,
        period_end_utc,
    )

    send_email(
        excel_data=excel_data,
        filename=filename,
        records=records,
        period_start_utc=period_start_utc,
        period_end_utc=period_end_utc,
    )

    print(
        f"Report byl odeslán na "
        f"{REPORT_RECIPIENT}."
    )

    print(
        f"Název přílohy: {filename}"
    )

    print(
        f"Počet záznamů: {len(records)}"
    )


if __name__ == "__main__":
    main()
